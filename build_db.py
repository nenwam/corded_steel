"""Build `cordedsteel.db` from the original spreadsheet.

Run once locally, then push the result to Turso:

    python3 build_db.py "Corded Steel 2026.xlsx"
    turso db create cordedsteel --from-file cordedsteel.db

Requires openpyxl (`pip install openpyxl`) — the app itself does not.
Re-running rebuilds the file from scratch, so don't run it against a database
that already holds real entries.
"""

from __future__ import annotations

import os
import sys
from datetime import datetime

import db

DEFAULT_XLSX = "Corded Steel 2026.xlsx"
DEFAULT_PASSWORD = "poop"

# Exercises whose values are fractional (miles run) rather than whole reps.
DECIMAL_EXERCISES = {"run"}
UNITS = {"run": "mi"}


def read_sheet(path):
    """Pull participants, exercises, goals and dates out of the tracker sheet.

    The sheet is laid out as a merged banner of person names above a row of
    exercise headers, with the goal for each column two rows further up:

        row 1   |        | Goal ..............| Goal ..............|
        row 2   |        | 3000 |  500 |  30  | 3000 |  500 |  30  |
        row 3   |        | Connor ............| Trevor ............|
        row 4   | Date   | Push-Ups | Pull-Ups | Run | Push-Ups | ...
        row 5+  | date   | values ...
    """
    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required to read the spreadsheet: pip install openpyxl")

    workbook = openpyxl.load_workbook(path, data_only=True)
    sheet = workbook.worksheets[0]

    header_row = next(
        (
            row
            for row in range(1, sheet.max_row + 1)
            if str(sheet.cell(row=row, column=1).value or "").strip().lower() == "date"
        ),
        None,
    )
    if header_row is None or header_row < 3:
        sys.exit("Could not find the 'Date' header row in the spreadsheet.")

    people_row = header_row - 1
    goal_row = header_row - 2

    # Merged banner cells only carry a value in their top-left cell, so the
    # person's name has to be carried forward across their exercise columns.
    columns = []
    current_person = None
    for col in range(2, sheet.max_column + 1):
        person = sheet.cell(row=people_row, column=col).value
        if person and str(person).strip():
            current_person = str(person).strip()
        exercise = sheet.cell(row=header_row, column=col).value
        if not current_person or not exercise or not str(exercise).strip():
            continue
        goal = sheet.cell(row=goal_row, column=col).value
        columns.append(
            {
                "column": col,
                "person": current_person,
                "exercise": str(exercise).strip(),
                "goal": float(goal) if isinstance(goal, (int, float)) else 0.0,
            }
        )

    rows = []
    for row in range(header_row + 1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=1).value
        if not isinstance(cell, datetime):
            break  # the Total / % of Goal / Avg. footer starts here
        values = {}
        for column in columns:
            value = sheet.cell(row=row, column=column["column"]).value
            if isinstance(value, (int, float)):
                values[(column["person"], column["exercise"])] = float(value)
        rows.append((cell.date(), values))

    if not rows:
        sys.exit("No dated rows found in the spreadsheet.")

    people, exercises = [], []
    for column in columns:
        if column["person"] not in people:
            people.append(column["person"])
        if column["exercise"] not in exercises:
            exercises.append(column["exercise"])

    goals = {(c["person"], c["exercise"]): c["goal"] for c in columns}
    return people, exercises, goals, rows


def build(path, password):
    people, exercises, goals, rows = read_sheet(path)

    target = db.local_path()
    if os.path.exists(target):
        os.remove(target)

    # Always the local file, never Turso — this script is destructive.
    conn = db.connect_local()
    db.init_schema(conn)

    db.set_password(conn, password)
    db.set_meta(conn, "title", "Corded Steel 2026")
    db.set_meta(conn, "start_date", rows[0][0].isoformat())
    db.set_meta(conn, "end_date", rows[-1][0].isoformat())

    exercise_ids = {}
    for position, name in enumerate(exercises):
        key = name.lower()
        db.execute(
            conn,
            "INSERT INTO exercises (name, unit, decimals, position) VALUES (?, ?, ?, ?)",
            (name, UNITS.get(key, ""), 2 if key in DECIMAL_EXERCISES else 0, position),
        )
        exercise_ids[name] = db.query(
            conn, "SELECT id FROM exercises WHERE name = ?", (name,)
        )[0][0]

    participant_ids = {}
    for position, name in enumerate(people):
        db.execute(
            conn,
            "INSERT INTO participants (name, position) VALUES (?, ?)",
            (name, position),
        )
        participant_ids[name] = db.query(
            conn, "SELECT id FROM participants WHERE name = ?", (name,)
        )[0][0]

    for (person, exercise), goal in goals.items():
        db.set_goal(conn, participant_ids[person], exercise_ids[exercise], goal)

    # Blank and zero cells are left out entirely — the app reads a missing cell
    # as 0, which keeps the table sparse and the sync to Turso small.
    cells = [
        (participant_ids[person], exercise_ids[exercise], day, value)
        for day, values in rows
        for (person, exercise), value in values.items()
        if value
    ]
    db.set_entries(conn, cells)

    # `turso db create --from-file` rejects anything that isn't in WAL mode, and
    # the checkpoint folds the -wal sidecar back into the file so the upload is
    # self-contained rather than missing the most recent writes.
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    conn.commit()
    conn.close()

    print(f"Wrote {target}")
    print(f"  participants : {', '.join(people)}")
    print(f"  exercises    : {', '.join(exercises)}")
    print(f"  dates        : {rows[0][0]} .. {rows[-1][0]} ({len(rows)} days)")
    print(f"  entries      : {len(cells)}")
    print(f"  password     : {password!r} (stored as a salted PBKDF2 hash)")


if __name__ == "__main__":
    source = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(source):
        sys.exit(f"Spreadsheet not found: {source}")
    build(source, os.environ.get("CORDED_STEEL_PASSWORD", DEFAULT_PASSWORD))
